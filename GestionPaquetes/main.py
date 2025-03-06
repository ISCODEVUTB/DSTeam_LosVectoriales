import openpyxl
import os
import hashlib
#gestion e
def hacer_pedido(usuario):
    print("\nLista de paquetes disponibles:")
    ExcelPaquetes.mostrar_dis()
    
    paquete = input("Ingrese el nombre del paquete que desea pedir: ")

    print(f"Pedido de '{paquete}' realizado con éxito.")

#funcion usuario
USUARIOS_FILE = "usuarios.txt"
cod_admin = [123, 321, 456, 654]

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def registrar_usuario():
    usuario = input("Ingrese el nombre de usuario: ")
    contraseña = input("Ingrese la contraseña: ")
    tipo = input("Ingrese el tipo de usuario (cliente, administrador, domiciliario): ").lower()

    if tipo not in ["cliente", "administrador", "domiciliario"]:
        print("Tipo de usuario no válido.")
        return
    
    if tipo == "administrador":
        codigo = int(input("Ingrese código de administrador: "))
        if codigo not in cod_admin:
            print("Código de administrador incorrecto.")
            return

    with open(USUARIOS_FILE, "a") as file:
        file.write(f"{usuario},{hash_password(contraseña)},{tipo}\n")
    print("Usuario registrado con éxito.")

def autenticar_usuario():
    usuario = input("Ingrese su nombre de usuario: ")
    contraseña = input("Ingrese su contraseña: ")

    with open(USUARIOS_FILE, "r") as file:
        for line in file:
            user, pass_hash, tipo = line.strip().split(",")
            if user == usuario and pass_hash == hash_password(contraseña):
                print(f"Autenticación exitosa. Bienvenido, {usuario} ({tipo}).")
                if tipo == "cliente":
                    menu_cliente(usuario)
                elif tipo == "administrador":
                    menu_administrador(usuario)
                elif tipo == "domiciliario":
                    menu_domiciliario(usuario)
                return  
    print("Usuario o contraseña incorrectos.")

def menu_cliente(usuario):
    while True:
        print(f"\nMenú Cliente ({usuario})")
        print("1. Tus productos")
        print("2. Hacer pedido")
        print("3. Ver estado de pedidos")
        print("4. Cerrar sesión")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            print("Mostrando productos...")
        elif opcion == "2":
            print("Realizando un pedido...")
            hacer_pedido(usuario)
        elif opcion == "3":
            print("Mostrando tus pedidos pendientes...")
        elif opcion == "4":
            print("Cerrando sesión...")
            break
        else:
            print("Opción no válida.")

def menu_administrador(usuario):
    while True:
        print(f"\nMenú Administrador ({usuario})")
        print("1. Agregar producto")
        print("2. Ver pedidos")
        print("3. Gestionar usuarios")
        print("4. Cerrar sesión")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            print("Agregando un nuevo producto...")
        elif opcion == "2":
            print("Listando pedidos...")
        elif opcion == "3":
            print("Gestionando usuarios...")
        elif opcion == "4":
            print("Cerrando sesión...")
            break
        else:
            print("Opción no válida.")

def menu_domiciliario(usuario):
    while True:
        print(f"\nMenú Domiciliario ({usuario})")
        print("1. Ver pedidos")
        print("2. Actualizar informacion de pedido")
        print("3. Cerrar sesión")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            print("Mostrando pedidos asignados...")
        elif opcion == "2":
            print("Pedido marcado como entregado.")
        elif opcion == "3":
            print("Cerrando sesión...")
            break
        else:
            print("Opción no válida.")

def menu():
    while True:
        print("\nMenú Principal")
        print("1. Registrar usuario")
        print("2. Iniciar sesión")
        print("3. Salir")
        opcion = input("Seleccione una opción: ")
        
        if opcion == "1":
            registrar_usuario()
        elif opcion == "2":
            autenticar_usuario()
        elif opcion == "3":
            print("Saliendo...")
            break
        else:
            print("Opción no válida.")

#gestion p
class Paquetes:
    def __init__(self, nombre, peso, precio, tipo, contenido=None, categoria=None, dimension=None, estado_pedido=None, creador = None, direccion = None, domiciliario = None):
        self.__nombre = nombre
        self.__peso = f"{peso} kg"
        self.__precio = f"{precio} $"
        self.__tipo = tipo
        self.__contenido = contenido if contenido else []
        self.__categoria = categoria if categoria else []
        self.__dimension = dimension if dimension else []
        self.__estado_pedido = estado_pedido
        self.__creador = creador
        self.__direccion = direccion
        self.__domiciliario = domiciliario


    @property
    def nombre(self):
        return self.__nombre
    
    @nombre.setter
    def nombre(self, new_name):
        self.__nombre = new_name

    @property
    def peso(self):
        return self.__peso
    
    @peso.setter
    def peso(self, new_peso):
        self.__peso = f"{new_peso} kg"

    @property
    def tipo(self):
        return self.__tipo

    @property
    def precio(self):
        return self.__precio

    @precio.setter
    def precio(self, new_price):
        self.__precio = f"{new_price} $"
    
    @tipo.setter
    def tipo(self, new_tipo):
        self.__tipo = new_tipo

    @property
    def contenido(self):
        return self.__contenido
    
    @contenido.setter
    def contenido(self, cont):
        self.__contenido.pop()
        self.__contenido.append(cont)

    @property
    def categoria(self):
        return self.__categoria
    
    @categoria.setter
    def categoria(self, cat):
        self.__categoria.pop()
        self.__categoria.append(cat)

    @property
    def dimension(self):
        return self.__dimension
    
    @dimension.setter
    def dimension(self, dim):
        self.__dimension.pop()
        self.__dimension.append(dim)

    @property
    def estado_pedido(self):
        return self.__estado_pedido
    
    def agregar_estado_pedido(self, nuevo_estado):
        self.__estado_pedido = nuevo_estado

    @property
    def direccion(self):
        return self.__direccion

    @direccion.setter
    def direccion(self, new_dir):
        self.__direccion = new_dir

    @property
    def domiciliario(self):
        return self.__domiciliario

    @domiciliario.setter
    def domiciliario(self, new_dom):
        self.__domiciliario = new_dom

    def __str__(self):
        return (f"Creador: {self.__creador}\n"
                f"Paquete: {self.__nombre}\n"
                f"precio: {self.__precio}\n"
                f"Peso: {self.__peso}\n"
                f"Tipo: {self.__tipo}\n"
                f"Contenido: {', '.join(self.__contenido) if self.__contenido else 'N/A'}\n"
                f"Categoría: {', '.join(self.__categoria) if self.__categoria else 'N/A'}\n"
                f"Dimensiones: {', '.join(self.__dimension) if self.__dimension else 'N/A'}\n"
                f"estado del pedido: {self.__estado_pedido}\n")
                
class Creacion:
    @staticmethod
    def crear_paquete(usuario):

        nombre = input("Nombre del paquete: ")
        precio= input("Digite el precio del paquete: ")
        peso = input("Peso (kg): ")
        tipo = None
        while tipo != "basico" or tipo != "estandar" or tipo != "dimensionado":
            print("El tipo debe ser basico, estandar o dimensionado") 
            tipo = input("Tipo (basico, estandar o dimensionado): ")
            if tipo == "basico" or tipo == "estandar" or tipo == "dimensionado":
                break
        contenido = input("Contenido (separado por comas): ").split(",") 
        categoria = input("Categorías (separadas por comas): ").split(",") 
        dimension = input("Dimensiones (ej: 10x20x30 cm): ").split(",")
        estado_pedido = "pendiente"
        direccion= input("Cual es la direccion de su domicilio? ")
        domiciliario= None
        return Paquetes(usuario, nombre, peso, precio, tipo, contenido, categoria, dimension, estado_pedido, direccion, domiciliario)

class ExcelPaquetes:
    FILE_PATH = "paquetes.xlsx"

    @classmethod
    def iniciar_excel(cls):
        if not os.path.exists(cls.FILE_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Creador", "Nombre", "Precio", "Peso", "Tipo", "Contenido", "Categoría", "Dimensiones", "estado pedido","Direccion","Domiciliario"])
            wb.save(cls.FILE_PATH)

    @classmethod
    def guardar(cls, paquete):
        if not os.path.exists(cls.FILE_PATH):
            cls.iniciar_excel()

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        ws.append([
            paquete.creador,
            paquete.nombre,
            paquete.precio,
            paquete.peso,
            paquete.tipo,
            ", ".join(paquete.contenido) if paquete.contenido else "N/A",
            ", ".join(paquete.categoria) if paquete.categoria else "N/A",
            ", ".join(paquete.dimension) if paquete.dimension else "N/A",
            paquete.estado_pedido,
            paquete.direccion,
            paquete.domiciliario,
            ])
        
        wb.save(cls.FILE_PATH)
        wb.close()
        print("\nPaquete guardado con éxito en la base de datos.")

    @classmethod
    def mostrar(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | {row[8]}" )
         
        wb.close()
    @classmethod    
    def mostrar_dis(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] == "pendiente":
                print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | Disponible ")
         
        wb.close()
    @classmethod
    def actualizar_datos(cls, usuarios):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        usuario = usuario 
        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active

        paquetes_usuario = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == usuario: 
                paquetes_usuario.append((row_idx, row))

        if not paquetes_usuario:
            print("No tienes paquetes registrados.")
            return

        print("\nTus productos:")
        for idx, (row_idx, paquete) in enumerate(paquetes_usuario, start=1):
            print(f"{idx}. {paquete[1]} | Precio: {paquete[2]} | Peso: {paquete[3]} | Tipo: {paquete[4]}")

        try:
            pcambio = int(input("\nSeleccione el número del paquete que desea cambiar: ")) - 1
            if pcambio < 0 or pcambio >= len(paquetes_usuario):
                print("Selección inválida.")
                return
        except ValueError:
            print("Entrada inválida. Debe ser un número.")
            return

        row_idx, paquete = paquetes_usuario[pcambio]

        print("\nSeleccione el dato que desea cambiar:")
        print(f"1. Nombre ({paquete[1]})")
        print(f"2. Precio ({paquete[2]})")
        print(f"3. Peso ({paquete[3]})")
        print(f"4. Tipo ({paquete[4]})")
        print(f"5. Contenido ({paquete[5]})")
        print(f"6. Categoría ({paquete[6]})")
        print(f"7. Dimensiones ({paquete[7]})")

        try:
            cambio = int(input("\nDigite el número de la opción a cambiar: "))
            if cambio < 1 or cambio > 7:
                print("Selección inválida.")
                return
        except ValueError:
            print("Entrada inválida. Debe ser un número.")
            return

        nuevo_valor = input("Ingrese el nuevo valor: ")

        ws.cell(row=row_idx, column=cambio + 1, value=nuevo_valor)

        wb.save(cls.FILE_PATH)
        wb.close()
        
        print("Cambio realizado con éxito.")

    @classmethod
    def registrar_pedido(cls):
        pass

ExcelPaquetes.iniciar_excel()
ExcelPaquetes.guardar()
#menu()