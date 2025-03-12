import openpyxl
import os

class ExcelPackages:
    FILE_PATH = "packages.xlsx"

    @classmethod
    def start_excel(cls, package):
        if not os.path.exists(cls.FILE_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([
                package.name, package.weight, package.price, package.type,
                package.content, package.category, package.dimension, package.order_status
            ])
            wb.save(cls.FILE_PATH)

    @classmethod
    def save(cls, package):
        if not os.path.exists(cls.FILE_PATH):
            cls.start_excel(package)

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        ws.append([
            package.name,
            package.price,
            package.weight,
            package.type,
            ", ".join(package.content) if package.content else "N/A",
            ", ".join(package.category) if package.category else "N/A",
            ", ".join(package.dimension) if package.dimension else "N/A",
            package.order_status,
            package.address,
            package.domiciliary,
        ])
        
        wb.save(cls.FILE_PATH)
        wb.close()
        print("\nPaquete guardado con éxito en la base de datos.")

    @classmethod
    def show(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | {row[8]}")
        
        wb.close()

    @classmethod    
    def show_dis(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] == "pendiente":
                print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | Disponible")
        
        wb.close()