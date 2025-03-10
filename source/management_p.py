import openpyxl
import os
op_n = "Opcion invalida"
class Packages:
    def __init__(self, name, weight, price, type, content=None, category=None, dimension=None, order_status=None, addreess = None, domiciliary = None):
        self.__name = name
        self.__weight = f"{weight} kg"
        self.__price = f"{price} $"
        self.__type = type
        self.__content = content if content else []
        self.__category = category if category else []
        self.__dimension = dimension if dimension else []
        self.__order_status = order_status
        self.__addreess = addreess
        self.__domiciliary = domiciliary


    @property
    def name(self):
        return self.__name
    
    @name.setter
    def name(self, new_name):
        self.__name = new_name

    @property
    def weight(self):
        return self.__weight
    
    @weight.setter
    def weight(self, new_weight):
        self.__weight = f"{new_weight} kg"

    @property
    def type(self):
        return self.__type

    @property
    def price(self):
        return self.__price

    @price.setter
    def price(self, new_price):
        self.__price = f"{new_price} $"
    
    @type.setter
    def type(self, new_type):
        self.__type = new_type

    @property
    def content(self):
        return self.__content
    
    @content.setter
    def content(self, cont):
        self.__content.pop()
        self.__content.append(cont)

    @property
    def category(self):
        return self.__category
    
    @category.setter
    def category(self, cat):
        self.__category.pop()
        self.__category.append(cat)

    @property
    def dimension(self):
        return self.__dimension
    
    @dimension.setter
    def dimension(self, dim):
        self.__dimension.pop()
        self.__dimension.append(dim)

    @property
    def order_status(self):
        return self.__order_status
    
    def agregar_order_status(self, nuevo_estado):
        self.__order_status = nuevo_estado

    @property
    def addreess(self):
        return self.__addreess

    @addreess.setter
    def addreess(self, new_addr):
        self.__addreess = new_addr

    @property
    def domiciliary(self):
        return self.__domiciliary

    @domiciliary.setter
    def domiciliary(self, new_dom):
        self.__domiciliary = new_dom

    def __str__(self):
        return (f"Paquete: {self.__name}\n"
                f"precio: {self.__price}\n"
                f"Peso: {self.__weight}\n"
                f"Tipo: {self.__type}\n"
                f"Contenido: {', '.join(self.__content) if self.__content else 'N/A'}\n"
                f"Categoría: {', '.join(self.__category) if self.__category else 'N/A'}\n"
                f"Dimensiones: {', '.join(self.__dimension) if self.__dimension else 'N/A'}\n"
                f"estado del pedido: {self.__order_status}\n")
                
class Creation:
    @staticmethod
    def create_packages():

        name = input("Nombre del paquete: ")
        price= input("Digite el precio del paquete: ")
        weight = input("Peso (kg): ")
        type = None
        while type != "basico" or type != "estandar" or type != "dimensionado":
            print("El tipo debe ser basico, estandar o dimensionado") 
            type = input("Tipo (basico, estandar o dimensionado): ")
            if type == "basico" or type == "estandar" or type == "dimensionado":
                break
        content = input("Contenido (separado por comas): ").split(",") 
        category = input("Categorías (separadas por comas): ").split(",") 
        dimension = input("Dimensiones (ej: 10x20x30 cm): ").split(",")
        order_status = "pendiente"
        addreess= None
        domiciliary= None
        return Packages(name, weight, price, type, content, category, dimension, order_status, addreess, domiciliary)

class ExcelPackages:
    FILE_PATH = "packages.xlsx"

    @classmethod
    def start_excel(cls):
        if not os.path.exists(cls.FILE_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([ "Nombre", "Precio", "Peso", "Tipo", "Contenido", "Categoría", "Dimensiones", "Estado pedido","Direccion","Domiciliario","Usuario"])
            wb.save(cls.FILE_PATH)

    @classmethod
    def save(cls, package):
        if not os.path.exists(cls.FILE_PATH):
            cls.start_excel()

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        ws.append([
            
            package.name,
            package.price,
            package.weight,
            package.type,
            ", ".join(package.content) if package.content else "N/A",
            ", ".join(package.category) if package.category else "N/A",
            ", ".join(package.dimension) if package.dimension else "N/A",
            package.order_status,
            package.addreess,
            package.domiciliary,
            ])
        
        wb.save(cls.FILE_PATH)
        wb.close()
        print("\nPaquete guardado con éxito en la base de datos.")

    @classmethod
    def show(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | {row[8]}" )
         
        wb.close()
    @classmethod    
    def show_dis(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] == "pendiente":
                print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | Disponible ")
         
        wb.close()


if __name__== "main":
    package = Creation.create_packages()  
    ExcelPackages.save(package) 
