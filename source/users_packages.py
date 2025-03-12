import hashlib
import openpyxl
import os

op_n, USERS_FILE, cod_admin = (
    "Opción inválida",
    "users.txt",
    {123, 321, 456, 654},
)


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def input_user_data():
    return input("Ingrese el nombre de usuario: "), input(
        "Ingrese la contraseña: "
    )


def register_user():
    user, pin = input_user_data()
    user_type = input(
        "Ingrese el tipo de usuario (cliente, administrador, domiciliario): "
    ).lower()
    if user_type not in {"cliente", "administrador", "domiciliario"}:
        return print("Tipo de usuario no válido.")

    if (
        user_type == "administrador"
        and int(input("Ingrese código de administrador: ")) not in cod_admin
    ):
        return print("Código de administrador incorrecto.")

    with open(USERS_FILE, "a") as file:
        file.write(f"{user},{hash_password(pin)},{user_type}\n")
    print("Usuario registrado con éxito.")


def authenticate_user():
    from menu_u import menu_customer, menu_administrator, menu_domiciliary

    user, pin = input_user_data()
    with open(USERS_FILE, "r") as file:
        for line in file:
            u, p_hash, user_type = line.strip().split(",")
            if user == u and p_hash == hash_password(pin):
                print(
                    f"Autenticación exitosa. Bienvenido, {user} ({user_type})."
                )
                {
                    "cliente": menu_customer,
                    "administrador": menu_administrator,
                    "domiciliario": menu_domiciliary,
                }[user_type](user)
                return
    print("Usuario o contraseña incorrectos.")


def update_d(cls, user):
    if (
        not os.path.exists(cls.FILE_PATH)
        or os.path.getsize(cls.FILE_PATH) == 0
    ):
        return print("No hay paquetes guardados.")

    wb = openpyxl.load_workbook(cls.FILE_PATH)

    if not wb.sheetnames:
        return print("El archivo de Excel no tiene hojas válidas.")

    ws = wb.active if wb.active else wb[wb.sheetnames[0]]

    if ws is None:
        return print("No se pudo obtener la hoja de Excel.")

    packages_user = [
        (i + 2, row)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True))
        if row and row[0] == user
    ]

    if not packages_user:
        return print("No tienes paquetes registrados.")

    print("\nTus productos:")
    for i, (_, p) in enumerate(packages_user, 1):
        print(f"{i}. {p[1]} | Precio: {p[2]} | Peso: {p[3]} | Tipo: {p[4]}")

    try:
        row_idx, package = packages_user[
            int(
                input("\nSeleccione el número del paquete que desea cambiar: ")
            )
            - 1
        ]
    except (ValueError, IndexError):
        return print("Selección inválida.")

    opciones = [
        "Nombre",
        "Precio",
        "Peso",
        "Tipo",
        "Contenido",
        "Categoría",
        "Dimensiones",
    ]
    print("\nSeleccione el dato que desea cambiar:")
    for i, op in enumerate(opciones, 1):
        print(f"{i}. {op} ({package[i]})")

    try:
        cambio = int(input("\nDigite el número de la opción a cambiar: ")) - 1
        if cambio not in range(7):
            raise ValueError
    except ValueError:
        return print(op_n)

    ws.cell(
        row=row_idx, column=cambio + 2, value=input("Ingrese el nuevo valor: ")
    )
    wb.save(cls.FILE_PATH)
    print("Cambio realizado con éxito.")
