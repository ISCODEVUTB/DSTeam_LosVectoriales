import hashlib
from source.management_p import ExcelPaquetes
from source.management_e import hacer_pedido
import openpyxl
import os
op_n = "Opcion invalida"
USUARIOS_FILE = "usuarios.txt"
cod_admin = [123, 321, 456, 654]

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def registrar_usuario():
    usuario = input("Ingrese el nombre de usuario: ")
    contraseña = input("Ingrese la contraseña: ")
    tipo = input("Ingrese el tipo de usuario (cliente, administrador, domiciliario): ").lower()

    if tipo not in ["cliente", "administrador", "domiciliario"]:
        print("Tipo de usuario no válido.")
        return
    
    if tipo == "administrador":
        codigo = int(input("Ingrese código de administrador: "))
        if codigo not in cod_admin:
            print("Código de administrador incorrecto.")
            return

    with open(USUARIOS_FILE, "a") as file:
        file.write(f"{usuario},{hash_password(contraseña)},{tipo}\n")
    print("Usuario registrado con éxito.")

def autenticar_usuario():
    usuario = input("Ingrese su nombre de usuario: ")
    contraseña = input("Ingrese su contraseña: ")

    with open(USUARIOS_FILE, "r") as file:
        for line in file:
            user, pass_hash, tipo = line.strip().split(",")
            if user == usuario and pass_hash == hash_password(contraseña):
                print(f"Autenticación exitosa. Bienvenido, {usuario} ({tipo}).")
                if tipo == "cliente":
                    menu_cliente(usuario)
                elif tipo == "administrador":
                    menu_administrador(usuario)
                elif tipo == "domiciliario":
                    menu_domiciliario(usuario)
                return  
    print("Usuario o contraseña incorrectos.")
    
def actualizar_datos(cls, usuario):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active

        paquetes_usuario = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == usuario: 
                paquetes_usuario.append((row_idx, row))

        if not paquetes_usuario:
            print("No tienes paquetes registrados.")
            return

        print("\nTus productos:")
        for idx, (row_idx, paquete) in enumerate(paquetes_usuario, start=1):
            print(f"{idx}. {paquete[1]} | Precio: {paquete[2]} | Peso: {paquete[3]} | Tipo: {paquete[4]}")

        try:
            pcambio = int(input("\nSeleccione el número del paquete que desea cambiar: ")) - 1
            if pcambio < 0 or pcambio >= len(paquetes_usuario):
                print("Selección inválida.")
                return
        except ValueError:
            print("Entrada inválida. Debe ser un número.")
            return

        row_idx, paquete = paquetes_usuario[pcambio]

        print("\nSeleccione el dato que desea cambiar:")
        print(f"1. Nombre ({paquete[1]})")
        print(f"2. Precio ({paquete[2]})")
        print(f"3. Peso ({paquete[3]})")
        print(f"4. Tipo ({paquete[4]})")
        print(f"5. Contenido ({paquete[5]})")
        print(f"6. Categoría ({paquete[6]})")
        print(f"7. Dimensiones ({paquete[7]})")

        try:
            cambio = int(input("\nDigite el número de la opción a cambiar: "))
            if cambio < 1 or cambio > 7:
                print(op_n)
                return
        except ValueError:
            print(op_n)
            return

        nuevo_valor = input("Ingrese el nuevo valor: ")

        ws.cell(row=row_idx, column=cambio + 1, value=nuevo_valor)

        wb.save(cls.FILE_PATH)
        wb.close()
        
        print("Cambio realizado con éxito.")

def menu_cliente(usuario):
    while True:
        print(f"\nMenú Cliente ({usuario})")
        print("1. Tus productos")
        print("2. Hacer pedido")
        print("3. Ver estado de pedidos")
        print("4.cambiar un pedido ")
        print("5. Cerrar sesión")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            print("Mostrando productos...")
        elif opcion == "2":
            print("Realizando un pedido...")
            hacer_pedido(usuario)
        elif opcion == "3":
            print("Mostrando tus pedidos pendientes...")
        elif opcion == "4" :
            actualizar_datos(ExcelPaquetes, usuario)  
        elif opcion == "5":
            print("Cerrando sesión...")
            break
        else:
            print(op_n)

def menu_administrador(usuario):
    while True:
        print(f"\nMenú Administrador ({usuario})")
        print("1. Agregar producto")
        print("2. Ver pedidos")
        print("3. Gestionar usuarios")
        print("4. Cerrar sesión")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            print("Agregando un nuevo producto...")
        elif opcion == "2":
            print("Listando pedidos...")
        elif opcion == "3":
            print("Gestionando usuarios...")
        elif opcion == "4":
            print("Cerrando sesión...")
            break
        else:
            print(op_n)

def menu_domiciliario(usuario):
    while True:
        print(f"\nMenú Domiciliario ({usuario})")
        print("1. Ver pedidos")
        print("2. Actualizar informacion de pedido")
        print("3. Cerrar sesión")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            print("Mostrando pedidos asignados...")
        elif opcion == "2":
            print("Pedido marcado como entregado.")
        elif opcion == "3":
            print("Cerrando sesión...")
            break
        else:
            print(op_n)

def menu():
    while True:
        print("\nMenú Principal")
        print("1. Registrar usuario")
        print("2. Iniciar sesión")
        print("3. Salir")
        opcion = input("Seleccione una opción: ")
        
        if opcion == "1":
            registrar_usuario()
        elif opcion == "2":
            autenticar_usuario()
        elif opcion == "3":
            print("Saliendo...")
            break
        else:
            print(op_n)

            
menu()